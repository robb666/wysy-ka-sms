from openpyxl import load_workbook
from datetime import date, timedelta
import datetime
import time
import re
import boto3

wb = load_workbook(filename="M:/Agent baza/2014 BAZA MAGRO.xlsx", read_only=True)
ws = wb['BAZA 2014']

# wb = load_workbook(filename="C:/Users/ROBERT/Desktop/TESTY.xlsx", read_only=True)
# ws = wb['Arkusz1']


today = date.today()
week_period = today - timedelta(-10)

def koniec_polisy():
    start_time = time.time()
    cells = ws['AE8000':'AF20000']
    lista = []
    for początek, koniec in cells:
        if koniec.value is not None and re.search('[0-9]', str(koniec.value)) and not re.search('[AWV()=.]',
                                                                                                str(koniec.value)):
            koniec_okresu = str(koniec.value)
            koniec_okresu_bez_sec = koniec_okresu[:10]
            if datetime.datetime.strptime(koniec_okresu_bez_sec, '%Y-%m-%d').date() == week_period:
                r = koniec.row
                if ws.cell(row=r, column=39).value != 'życ' and ws.cell(row=r, column=48).value is not None:

                    rozlicz = ws.cell(row=r, column=7).value
                    if rozlicz == 'Filipiak':
                        rozlicz = 'Ultimatum, tel 694888197'
                    elif rozlicz == 'Pankiewicz':
                        rozlicz = 'R. Pankiewiczem, tel 577839889'
                    elif rozlicz == 'Wawrzyniak':
                        rozlicz = 'A. Wawrzyniak, tel 691602675'
                    elif rozlicz == 'Wołowski':
                        rozlicz = 'M. Wołowskim, tel 692830084'
                    elif rozlicz == 'Robert':
                        rozlicz = 'MAGRO, tel 572810576'
                    else:
                        rozlicz = 'MAGRO, tel 602752893'

                    przedmiot_ub = ws.cell(row=r, column=25).value
                    if przedmiot_ub is None:
                        przedmiot_ub = ''
                    nr_polisy = ws.cell(row=r, column=40).value


                    nr_tel = ws.cell(row=r, column=19).value
                    tel = str(nr_tel)
                    if re.search('^[4]', tel):  # numer domowy
                        tel = ''
                    if re.search(r'[0-9]', tel):
                        tel = '48' + tel.replace(' ', '').strip('+')
                        if re.search('[a-zA-z;:?,]', tel):
                            tel = tel[:11]
                        if len(tel) > 11:
                            tel = tel[2:13]

                        lista.append(tel)
                        lista.append(koniec_okresu_bez_sec)
                        lista.append(nr_polisy)
                        lista.append(przedmiot_ub)
                        lista.append(rozlicz)

                    print()
                else:
                    print('życie k...')

    return start_time, lista


def wysyłka_aws(start_time, lista):
    # print(lista)
    # list(dict.fromkeys([i for i in lista[::5]])) # --usuwanie duplikatów z listy--

    tel = [i for i in lista[::5]]
    koniec_okresu_bez_sec = [j for j in lista[1::5]]
    nr_polisy = [k for k in lista[2::5]]
    przedmiot_ub = [l for l in lista[3::5]]
    rozlicz = [l for l in lista[4::5]]

    n = 0
    for m in tel:
        # print(m)

        client = boto3.client('sns', 'eu-west-1')

        client.publish(PhoneNumber=str(tel[n]), Message='Dnia ' + koniec_okresu_bez_sec[n] +
                                                        ' dobiega końca Twoja polisa ubezpieczeniowa, nr. ' +
                                                        nr_polisy[n] + ', ' + przedmiot_ub[n] +
                                                        '. W spr odnowienia prosimy o kontakt z ' + rozlicz[n])

        print(str(tel[n]) + ' - Dnia ' + koniec_okresu_bez_sec[n] + ' dobiega końca Twoja polisa ubezpieczeniowa, nr. '
              + nr_polisy[n] + ', ' + przedmiot_ub[n] + '. W spr odnowienia prosimy o kontakt z ' + rozlicz[n])

        print()
        n += 1

    end_time = time.time() - start_time
    print()
    print()
    print('Czas wykonania: {:.4f} minut'.format(end_time / 60))
    time.sleep(120)

def main():

    start_time, lista = koniec_polisy()
    wysyłka_aws(start_time, lista)

main()
