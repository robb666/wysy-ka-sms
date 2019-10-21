from openpyxl import load_workbook
from datetime import date, timedelta
import datetime
import time
import re
import boto3

wb = load_workbook(filename="M:/Agent baza/DO wznowienia.xlsx", read_only=True)
ws = wb['do oferty']

# wb = load_workbook(filename="C:/Users/ROBERT/Desktop/TESTY.xlsx", read_only=True)
# ws = wb['Arkusz1']

today = date.today()
week_period = today - timedelta(-10)

def koniec_okresu():
    start_time = time.time()
    cells = ws['AE1':'AF2000']
    lista = []
    for początek, koniec in cells:
        if koniec.value is not None and re.search('[0-9]', str(koniec.value)) and not re.search('[AWV()=.]',
                                                                                                str(koniec.value)):
            koniec_okresu = str(koniec.value)
            koniec_okresu_bez_sec = koniec_okresu[:10]
            if datetime.datetime.strptime(koniec_okresu_bez_sec, '%Y-%m-%d').date() == week_period:
                r = koniec.row

                nr_tel = ws.cell(row=r, column=19).value
                tel = str(nr_tel)
                if re.search('^[4]', tel):  # numer domowy
                    tel = ''
                if re.search(r'[0-9]', tel):
                    tel = '48' + tel.replace(' ', '').strip('+')
                    if re.search('[a-zA-z;:?,]', tel):
                        tel = tel[:11]
                    if len(tel) > 11:
                        tel = tel[2:13]

                marka_kod = ws.cell(row=r, column=23).value
                if marka_kod is None:
                    marka_kod = ''
                model_poczta = ws.cell(row=r, column=24).value
                if model_poczta is None:
                    model_poczta = ''
                przedmiot_ub = ws.cell(row=r, column=25).value
                if przedmiot_ub is None:
                    przedmiot_ub = ''

                lista.append(tel)
                lista.append(marka_kod)
                lista.append(model_poczta)
                lista.append(przedmiot_ub)

    return lista


def wysyłka_aws(lista):

    tel = [i for i in lista[::4]]
    marka_kod = [j for j in lista[1::4]]
    model_poczta = [k for k in lista[2::4]]
    przedmiot_ub = [l for l in lista[3::4]]

    n = 0
    for m in tel:
        client = boto3.client('sns', 'eu-west-1')

        client.publish(PhoneNumber=str(tel[n]), Message='Zbliża się koniec ubezpieczenia ' + marka_kod[n] + ' ' \
                                                           + model_poczta[n] + ', ' + przedmiot_ub[n] + \
                                                           '. W celu otrzymania najlepszej oferty prosimy o kontakt z' \
                                                           'naszą Agencją - MAGRO Ubezpieczenia, tel. 572810576\n' \
                                                           'https://ubezpieczenia-magro.pl')

        print()
        text = 'Zbliża się koniec ubezpieczenia ' + marka_kod[n] + ' ' \
                                                           + model_poczta[n] + ', ' + przedmiot_ub[n] + \
                                                           '. W celu otrzymania najlepszej oferty prosimy o kontakt z' \
                                                           'naszą Agencją - MAGRO Ubezpieczenia, tel. 572810576\n' \
                                                           'https://ubezpieczenia-magro.pl'

        print(text)
        n += 1
    time.sleep(60)

def main():
    lista = koniec_okresu()
    wysyłka_aws(lista)


main()