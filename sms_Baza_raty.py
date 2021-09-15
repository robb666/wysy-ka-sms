from openpyxl import load_workbook
from datetime import date, datetime, timedelta
import re
import boto3
import time

start_time = time.time()
now = datetime.now().strftime("Wysłane dnia %d.%m.%Y o godzinie %H:%M:%S")
print(f'Wysyłka SMS - Przypomnienia o ratach.\n{now}')

# wb = load_workbook(filename="M:/Agent baza/2014 BAZA MAGRO.xlsx", read_only=False, data_only=True)
wb = load_workbook(filename="/run/user/1000/gvfs/smb-share:server=192.168.1.12,share=e/Agent baza/2014 BAZA MAGRO.xlsx",
                   read_only=False, data_only=True)
ws = wb['BAZA 2014']

today = date.today()
week_period = today - timedelta(-5)

print('SMS - Przypomnienia o ratach.')

def dane_rat():

    # cells_signed = ws['G4178':f'H{ws.max_row}']
    cells = ws['AW4178':f'BA{ws.max_row}']
    # for podpis, H in cells_signed:

        # if podpis.value == 'MAGRO' or podpis.value == 'Robert':
        #     print(podpis.value)

    list = []

    for data_raty, kwota, x, y, nr_raty in cells:
        if data_raty.value is not None and re.search('[0-9]', str(data_raty.value)) and \
                                                                    not re.search('[AWV()=.]', str(data_raty.value)):
            data_r = str(data_raty.value)
            termin_płatności = data_r[:10]
            if datetime.strptime(termin_płatności, '%Y-%m-%d').date() == week_period and \
                nr_raty.value is not None and nr_raty.value > 1:

                r = data_raty.row
                kwota_raty = kwota.value
                rodzaj = ws.cell(row=r, column=39).value
                if rodzaj in 'życ':
                    pass
                else:
                    nr_polisy = ws.cell(row=r, column=40).value
                    nr_tel = ws.cell(row=r, column=19).value
                    tel = str(nr_tel)
                    if tel.startswith('42'):  # numer domowy
                        tel = ''
                    if re.search(r'[0-9]', tel):
                        tel = '48' + tel.replace(' ', '').strip('+')
                        if re.search('[a-zA-z;:?,]', tel):
                            tel = tel[:11]
                        if len(tel) > 11:
                            tel = tel[2:13]

                        list.append(tel)
                        list.append(nr_polisy)
                        list.append(termin_płatności)
                        list.append(kwota_raty)

    return list


def wysyłka_aws(list):
    tel = [i for i in list[::4]]
    nr_polisy = [j for j in list[1::4]]
    termin_płatności = [k for k in list[2::4]]
    kwota_raty = [l for l in list[3::4]]

    n = 0
    for _ in termin_płatności:
        client = boto3.client('sns', 'eu-west-1')

        client.publish(PhoneNumber=str(tel[n]), Message='Przypomnienie o płatności raty: ' + str(kwota_raty[n]) +
                                                        ' zł, za polisę nr. ' + str(nr_polisy[n]) +
                                                        ' upływającym dnia ' + str(termin_płatności[n]) +
                                                        '. \n\nubezpieczenia-magro.pl/kalkulatorOC')

        print(str(tel[n]), 'Przypomnienie o płatności raty: ' + str(kwota_raty[n]) +
                                                    ' zł, za polisę nr. ' + str(nr_polisy[n]) +
                                              ' upływającym dnia ' + str(termin_płatności[n]) +
                                                   '. \n\nubezpieczenia-magro.pl/kalkulatorOC')

        print()
        n += 1


def main():
    list = dane_rat()
    wysyłka_aws(list)


main()

end_time = (time.time() - start_time)

print('\nCzas wykonania: ' + '{:.2f} sek'.format(end_time))
print('\n\n')
print('______________________________________')

# time.sleep(10)
