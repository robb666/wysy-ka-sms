from openpyxl import load_workbook
from datetime import date, timedelta
import datetime
import re
import boto3

wb = load_workbook(filename="M:/Agent baza/2014 BAZA MAGRO.xlsx", read_only=True)
ws = wb['BAZA 2014']

today = date.today()
week_period = today - timedelta(-3)

def dane_rat():
    cells = ws['AW8000':'BA20000']

    for data_raty, kwota, x, y, nr_raty in cells:

        if data_raty.value is not None and re.search('[0-9]', str(data_raty.value)) and not re.search('[AWV()=.]', str(data_raty.value)):
            data_r = str(data_raty.value)
            termin_płatności = data_r[:10]

            if datetime.datetime.strptime(termin_płatności, '%Y-%m-%d').date() == week_period and int(nr_raty.value) > 1:
                r = data_raty.row
                kwota_raty = kwota.value
                nr_polisy = ws.cell(row=r, column=40).value
                nr_tel = ws.cell(row=r, column=19).value
                tel = str(nr_tel)
                if re.search('^[4]', tel):     # numer domowy
                    tel = ''
                if re.search(r'[0-9]', tel):
                    tel = '48' + tel.replace(' ', '').strip('+')
                    if re.search('[a-zA-z;:?,]', tel):
                        tel = tel[:11]
                    if len(tel) > 11:
                        tel = tel[2:13]


                    print(tel)
                    print(nr_polisy)
                    print(termin_płatności)
                    print('{} {}'.format(kwota_raty, 'zł'))
                    print()



                    client = boto3.client('sns', 'eu-west-1')

                    client.publish(PhoneNumber=str(tel), Message='Przypomnienie o terminie raty: ' + str(kwota_raty) +
                                                                  ' zł, za polisę nr. ' + str(nr_polisy) +
                                                                   ' upływającym dnia ' + str(termin_płatności) +
                                                                    '. https://ubezpieczenia-magro.pl')



dane_rat()









