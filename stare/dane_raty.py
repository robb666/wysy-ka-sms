from openpyxl import load_workbook
from datetime import date, timedelta
import datetime
import re

wb = load_workbook(filename="M:/Agent baza/2014 BAZA MAGRO.xlsx", read_only=True)
ws = wb['BAZA 2014']


today = date.today()
week_period = today - timedelta(-8)



def data_raty():
    cells = ws['AW8000':'BA20000']
    for data_raty, kwota, x, y, nr_raty in cells:
        if data_raty.value is not None and \
                re.search('[0-9]', str(data_raty.value)) and not re.search('[AWV()=.]', str(data_raty.value)):
            data_r = str(data_raty.value)
            termin_płatności = data_r[:10]
            if datetime.datetime.strptime(termin_płatności, '%Y-%m-%d').date() == week_period and\
                    int(nr_raty.value) > 1:
                r = data_raty.row
                kwota_raty = kwota.value
                nr_polisy = ws.cell(row=r, column=40).value
                nr_tel = ws.cell(row=r, column=19).value
                tel = str(nr_tel)
                if re.search('^[4]', tel):
                    tel = ''
                if re.search(r'[0-9]', tel):
                    tel = '48' + tel.replace(' ', '').strip('+')
                    if re.search('[a-zA-z;:?,]', tel):
                        tel = tel[:11]
                    if len(tel) > 11:
                        tel = tel[2:13]

                    return kwota_raty, termin_płatności, nr_polisy, tel
























                # if re.search('^[4]', nr_tel):
                #     nr_tel = ''
                #     print('nr domowy')
                # if re.search('[0-9| ]', nr_tel):
                #     nr_tel = '48' + nr_tel.replace(' ', '')
                #
                #     if re.search('[a-zA-z;]', nr_tel):
                #         nr_tel = nr_tel[:11]
                #
                #     # else:
                #     #     pass
                #
                #     print(nr_tel)














# def nr_tel():
#     cells = ws['S3':'T20000']
#     for tel, mail in cells:
#         tel = str(tel.value)
#         if tel is not None:
#             if re.search('^[4]', tel):
#                 tel = ''
#                 # print('nr domowy')
#             if re.search(r'[0-9]', tel):
#                 tel = '48' + tel.replace(' ', '').strip('+')
#
#
#                 if re.search('[a-zA-z;:?,]', tel):
#                     tel = tel[:11]
#                 if len(tel) > 11:
#                     tel = tel[2:13]
#
#                 print(tel)









