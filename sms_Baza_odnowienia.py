from openpyxl import load_workbook
from datetime import date, timedelta
import datetime
import time
import re
import boto3

start_time = time.time()

print('Wysyłka SMS - Przypomnienia o odnowieniach.')


class SMS:

    def __init__(self):
        # self.wb = load_workbook(filename="M:/Agent baza/2014 BAZA MAGRO.xlsx", read_only=True)
        self.wb = load_workbook(filename="/run/user/1000/gvfs/smb-share:server=10.0.0.4,share=e/Agent baza/2014 BAZA MAGRO.xlsx",
                                read_only=True)
        self.ws = self.wb['BAZA 2014']
        self.cells = self.ws['G4178':f'AV{self.ws.max_row}']
        today = date.today()
        self.week_period = today - timedelta(-9)

    def read_excel(self):
        for rozlicz, H, I, J, K, L, M, N, O, P, Q, R, nr_tel, email, U, V, marka, model, przedmiot_ub, rok_prod, SU, \
                AB, AC, AD, pocz, koniec, AG, AH, AI, AJ, AK, tu, rodz_ub, nr_polisy, AO, AP, AQ, AR, AS, AT, AU, \
                    przypis in self.cells:
            self.rozlicz = rozlicz.value
            self.nr_tel = nr_tel.value
            self.przedmiot_ub = przedmiot_ub.value
            if self.przedmiot_ub is None:
                self.przedmiot_ub = ''
            self.tu = tu.value
            self.koniec = koniec.value
            self.rodz_ub = rodz_ub.value
            self.nr_polisy = nr_polisy.value
            self.przypis = przypis.value

            yield self.koniec

    def select_cells(self):
        for self.koniec in self.read_excel():
            if self.koniec is not None and re.search('[0-9]', str(self.koniec)) and not \
                    re.search('[AWV()=.]', str(self.koniec)):

                koniec_okresu = str(self.koniec)
                self.koniec_okresu_bez_sec = koniec_okresu[:10]
                if datetime.datetime.strptime(str(self.koniec_okresu_bez_sec), '%Y-%m-%d').date() == self.week_period:
                    if self.nr_tel is not None and self.rodz_ub != 'życ' and self.przypis is not None:
                        d = {'Filipiak': 'Ultimatum, tel. 694888197', 'Nowakowski': 'K. Nowakowskim, tel. 508280760',
                             'Pankiewicz': 'R. Pankiewiczem, tel. 577839889', 'Skrzypek': 'S. Skrzypkiem, tel. 508280764',
                             'Wawrzyniak': 'A. Wawrzyniak, tel. 691602675', 'Wołowski': 'M. Wołowskim, tel. 692830084',
                             'MAGRO': 'MAGRO, tel. 602752893', 'Robert': 'MAGRO, tel. 572810576'}
                        if self.rozlicz in d:
                            self.rozlicz = d.get(self.rozlicz)

                        di = {'ALL': 'Allianz',
                              'AXA': 'AXA',
                              'COM': 'Compensa',
                              'EIN': 'Euroins',
                              'EPZU': 'PZU',
                              'GEN': 'Generali',
                              'ŻGEN': 'Generali',
                              'GOT': 'Gothaer',
                              'HDI': 'HDI',
                              'HES': 'Ergo Hestia',
                              'IGS': 'IGS',
                              'INT': 'INTER',
                              'LIN': 'LINK 4',
                              'MTU': 'MTU',
                              'PRO': 'Proama',
                              'PZU': 'PZU',
                              'RIS': 'InterRisk',
                              'TUW': 'TUW', 'TUZ': 'TUZ',
                              'UNI': 'Uniqa',
                              'WAR': 'Warta',
                              'ŻWAR': 'Warta',
                              'WIE': 'Wiener',
                              'YCD': 'You Can Drive',
                              'None': ''}

                        self.tu = di.get(self.tu)
                        self.nr_tel = str(self.nr_tel)
                        if self.nr_tel.startswith('42'):  # numer domowy
                            self.nr_tel = ''
                        if re.search(r'[0-9]', self.nr_tel):
                            self.nr_tel = '48' + self.nr_tel.replace(' ', '').strip('+')
                            if re.search('[a-zA-z;:?,]', self.nr_tel):
                                self.nr_tel = self.nr_tel[:11]
                            if len(self.nr_tel) > 11:
                                self.nr_tel = self.nr_tel[2:13]
                        # print(self.koniec_okresu_bez_sec)
                        yield self.koniec_okresu_bez_sec

    def wysyłka_aws(self):
        for _ in self.select_cells():
            client = boto3.client('sns', 'eu-west-1')
            client.publish(PhoneNumber=str(self.nr_tel), Message='Dnia ' + str(self.koniec_okresu_bez_sec)
                                                        + ' dobiega końca Twoja polisa ubezpieczeniowa, nr. '
                                                        + str(self.nr_polisy) + ' - ' + str(self.tu) + ', '
                                                        + str(self.przedmiot_ub)
                                                        + '. W spr odnowienia prosimy o kontakt z ' + str(self.rozlicz)
                                                        + '\n\nubezpieczenia-magro.pl/kalkulatorOC')

            print(str(self.nr_tel + ' - Dnia ' + self.koniec_okresu_bez_sec
                      + ' dobiega końca Twoja polisa ubezpieczeniowa, nr. '
                      + self.nr_polisy + ' - ' + self.tu + ', '
                      + self.przedmiot_ub + '. W spr odnowienia prosimy o kontakt z ' + self.rozlicz
                      + '\n\nubezpieczenia-magro.pl/kalkulatorOC'))
            print()


odnowienia = SMS()
odnowienia.read_excel()
odnowienia.select_cells()
odnowienia.wysyłka_aws()

end_time = time.time() - start_time
print()
print('Czas wykonania: {:.0f} sekund'.format(end_time))
print('__________________________________')
print()
print()
# time.sleep(12)

