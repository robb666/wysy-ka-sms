from openpyxl import load_workbook
from datetime import date, timedelta
import datetime
import time
import re
import boto3

wb = load_workbook(filename="C:/Users/ROBERT/Desktop/TESTY.xlsx", read_only=True)
ws = wb['Arkusz1']

today = date.today()
week_period = today - timedelta(2)


def dane_rat():
    start_time = time.time()
    cells = ws['AW1':'BA10']
    list = []
    for data_raty, kwota, x, y, nr_raty in cells:
        if data_raty.value is not None and re.search('[0-9]', str(data_raty.value)) and not re.search('[AWV()=.]', str(data_raty.value)):
            data_r = str(data_raty.value)
            termin_płatności = data_r[:10]
            if datetime.datetime.strptime(termin_płatności, '%Y-%m-%d').date() == week_period and int(nr_raty.value) > 1:
                r = data_raty.row
                kwota_raty = kwota.value
                nr_polisy = ws.cell(row=r, column=40).value
                nr_tel = ws.cell(row=r, column=19).value
                tel = str(nr_tel)
                if re.search('^[4]', tel):     # numer domowy
                    tel = ''
                if re.search(r'[0-9]', tel):
                    tel = '48' + tel.replace(' ', '').strip('+')
                    if re.search('[a-zA-z;:?,]', tel):
                        tel = tel[:11]
                    if len(tel) > 11:
                        tel = tel[2:13]

                    list.append(tel)
                    list.append(nr_polisy)
                    list.append(termin_płatności)
                    list.append(kwota_raty)

    return start_time, list


def wysyłka_aws(start_time, list):

    tel = [i for i in list[::4]]
    nr_polisy = [j for j in list[1::4]]
    termin_płatności = [k for k in list[2::4]]
    kwota_raty = [l for l in list[3::4]]

    n = 0
    for m in termin_płatności:

        # client = boto3.client('sns', 'eu-west-1')
        #
        # client.publish(PhoneNumber=str(tel[n]), Message='Przypomnienie o płatności raty: ' + str(kwota_raty[n]) +
        #                                               ' zł, za polisę nr. ' + str(nr_polisy[n]) +
        #                                                ' upływającym dnia ' + str(termin_płatności[n]) +
        #                                                 '. https://ubezpieczenia-magro.pl')

        print(str(tel[n]), 'Przypomnienie o płatności raty: ' + str(kwota_raty[n]) +
                                                    ' zł, za polisę nr. ' + str(nr_polisy[n]) +
                                              ' upływającym dnia ' + str(termin_płatności[n]) +
                                                   '. https://ubezpieczenia-magro.pl')


        print()
        n += 1

        end_time = time.time() - start_time
        print(end_time, 'sekund')
        time.sleep(120)


def main():

    start_time, list = dane_rat()
    wysyłka_aws(start_time, list)



main()










# print(str(tel[n]), 'Przypomnienie o płatności raty: ' + str(kwota_raty[n]) +
#                                             ' zł, za polisę nr. ' + str(nr_polisy[n]) +
#                                       ' upływającym dnia ' + str(termin_płatności[n]) +
#                                            '. https://ubezpieczenia-magro.pl')